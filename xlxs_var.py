# -*- coding: utf8 -*-
import telebot
import openpyxl
from openpyxl.utils import get_column_letter
import re
from telebot import types
import datetime


bot = telebot.TeleBot('TOKEN')

def RASPISANIE(number, day, Name_day):
    raspisanie = []
    path_to_file = 'timesheet.xlsx'
    search_text = number.upper()
    wb = openpyxl.load_workbook(path_to_file)  # Грузим наш прайс-лист
    sheets_list = wb.sheetnames  # Получаем список всех листов в файле
    sheet_active = wb[sheets_list[day]]  # Начинаем работать с самым первым
    row_max = sheet_active.max_row  # Получаем количество столбцов
    # print(type(row_max))
    column_max = sheet_active.max_column  # Получаем количество строк
    row_min = 1  # Переменная, отвечающая за номер строки
    column_min = 1  # Переменная, отвечающая за номер столбца
    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)
            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min
            data_from_cell = sheet_active[word_cell].value
            data_from_cell = str(data_from_cell)
            regular = search_text
            result = re.findall(regular, data_from_cell)
            if len(result) > 0:
                cell = word_column
            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1
    sheet = wb.active
    para = ["1 Пара",
            "2 Пара",
            "3 Пара",
            "4 Пара",
            "5 Пара",
            "6 Пара",
            "7 Пара"]
    calls = ["8.00-9.20",
             "9.30-10.50",
             "11.10-12.30",
             "12.50-14.10",
             "14.30-15.50",
             "16.10-17.30",
             "17.40-19.00"]
    diapaz = [4, 5, 6, 7, 8, 9, 10]

    raspisanie.append(f'Группа - {search_text}')
    raspisanie.append(f'✅{Name_day}✅')
    for i in range(0, 7):
        if sheet[cell + str(diapaz[i])].value == None:
            sheet[cell + str(diapaz[i])].value = "Пары нет"
        raspisanie.append(f'⏰{calls[i]}: {para[i]} - 📕{sheet[cell + str(diapaz[i])].value}')
    return raspisanie

def WEEKEND_today(number, Name_day):
    search_text = number.upper()
    raspisanie = []
    raspisanie.append(f'Группа - {search_text}')
    raspisanie.append(f'✅{Name_day}✅')
    raspisanie.append(f'Сегодня можно отдохнуть 😴 🍻')
    return raspisanie
def WEEKEND_tommorow(number, Name_day):
    search_text = number.upper()
    raspisanie = []
    raspisanie.append(f'Группа - {search_text}')
    raspisanie.append(f'✅{Name_day}✅')
    raspisanie.append(f'Завтра можно отдохнуть 😴 🍻')
    return raspisanie
# -----------------------------------------------------------------------------------------------------------------------
number = ''
day = ''
group_list = ['11ГД', '11УМД', '11СХ', '21ГД', '21ГСк', '22ГД', '31ГД', '31ГСк',
              '32ГД', '41ГД', '42ГД', '21М', '31М', '32М', '42М', '21СХ', '31СХ', '21КЛ',
              '21П', '11КЛ', '11П', '11ИС', '11ММР', '11РМ', '12РМ', '22 ИС', '21ИС', '21РМ',
              '22РМ', '21КСК', '31ПИ', '32ПИ', '31РМ', '32РМ', '31КСК', '41КСК', '41ПИ', '42ПИ'
              ]

def keys(message):
    keyboard = telebot.types.ReplyKeyboardMarkup(True)
    keyboard.row('Расписание', 'Выбор группы')
    keyboard.row('Помощь')
    keyboard.row('/start')
    bot.send_message(message.chat.id, 'Привет  {}! \n'
                                      'Если ты учишься в КИГМ №23, то этот бот сможет показывать тебе твое расписание.\n'
                                      'Выбрать группу - нажмите /group.\n'
                                      'Чтобы получить расписание /timetable.\n'
                                      'Если возникили проблемы - нажмите /help'.format(message.chat.first_name),
                                      reply_markup=keyboard
                     )




@bot.message_handler(commands=['start'])
def start_command(message):
    keys(message)

@bot.message_handler(commands=['help'])
def help_command(message):
    keyboard = telebot.types.InlineKeyboardMarkup()
    keyboard.add(
        telebot.types.InlineKeyboardButton(
            'Написать разработчику', url='telegram.me/<YOUR TG>'
  )
    )
    bot.send_message(
        message.chat.id,
        '1) Напишите свою группу (не важно большими или маленькими буквами).\n' +
        '2) Выберете нужный вам день недели.\n' +
        '3) Вы получите расписание для выбранной вами группы и выбранного дня недели.\n' +
        '4) Чтобы получить расписание на другой день, просто нажмите на нужный вам день выше.',
        reply_markup=keyboard
    )
def error(message):
    bot.send_message(message.chat.id, 'Вы не выбрали группу /group\n'
                                      '\n'
                                      '\n'
                                      '/help\n')


@bot.message_handler(commands=['group'])
def group_command(message):
    group_com = bot.send_message(message.from_user.id, "{}, из какой вы группы?".format(message.chat.first_name))
    bot.register_next_step_handler(group_com, number_group)
def number_group(message):
    global number
    number = message.text
    if message.text == '/start':
        start_command(message)
    elif message.text.upper() in group_list:
        msg = (message.text).upper()
        group = (f'Выбрана группа: {msg}')
        day_number(message)
    else:
        bot.send_message(message.from_user.id, 'Увы, но такой группы нет, введите группу правильно')
        group_command(message)

@bot.message_handler(commands=['timetable'])
def day_number(message):
    if number == '':
        error(message)
    else:
        keyboard = types.InlineKeyboardMarkup()
        key_monday = types.InlineKeyboardButton(text='Понедельник', callback_data='monday')
        keyboard.add(key_monday)

        key_tuesday = types.InlineKeyboardButton(text='Вторник', callback_data='tuesday')
        keyboard.add(key_tuesday)

        key_wednesday = types.InlineKeyboardButton(text='Среда', callback_data='wednesday')
        keyboard.add(key_wednesday)

        key_thursday = types.InlineKeyboardButton(text='Четверг', callback_data='thursday')
        keyboard.add(key_thursday)

        key_friday = types.InlineKeyboardButton(text='Пятница', callback_data='friday')
        keyboard.add(key_friday)

        key_today = types.InlineKeyboardButton(text='Сегодня', callback_data='today')
        key_tomorrow = types.InlineKeyboardButton(text='Завтра', callback_data='tomorrow')
        keyboard.add(key_today, key_tomorrow)

        bot.send_message(message.chat.id, 'Выбери день недели', reply_markup=keyboard)


        rasp_sms = bot.send_message(message.chat.id, 'Выберете день и расписнаие поятвится тут')
        global id_sms
        id_sms = rasp_sms.id
def msg_out(message,mes):
    message_output = bot.edit_message_text(chat_id=message.chat.id, message_id=id_sms, text='\n'.join(mes))

@bot.callback_query_handler(func=lambda call: True)
def callback_worker(call):
    weekDays = ("Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскреенье")
    now = datetime.date.today()
    tomorrow = datetime.date.today() + datetime.timedelta(days=1)

    if call.data == "today":
        day = now.weekday()
        Name_day = weekDays[day]
        if day == 5:
            mes = WEEKEND_today(number, Name_day)
            msg_out(call.message, mes)
        elif day == 6:
            mes = WEEKEND_today(number, Name_day)
            msg_out(call.message, mes)
        else:
            mes = RASPISANIE(number, day, Name_day)
            msg_out(call.message, mes)
    elif call.data == "tomorrow":
        day = int(tomorrow.weekday())
        Name_day = weekDays[day]
        if day == 5:
            mes = WEEKEND_tommorow(number, Name_day)
            msg_out(call.message, mes)
        elif day == 6:
            mes = WEEKEND_tommorow(number, Name_day)
            msg_out(call.message, mes)
        else:
            mes = RASPISANIE(number, day, Name_day)
            msg_out(call.message, mes)
    elif call.data == "monday":
        day = 0
        Name_day = 'Понедельник'
        mes = RASPISANIE(number, day, Name_day)
        msg_out(call.message, mes)
    elif call.data == "tuesday":
        day = 1
        Name_day = 'Вторник'
        mes = RASPISANIE(number, day, Name_day)
        msg_out(call.message, mes)
    elif call.data == "wednesday":
        day = 2
        Name_day = 'Среда'
        mes = RASPISANIE(number, day, Name_day)
        msg_out(call.message, mes)
    elif call.data == "thursday":
        day = 3
        Name_day = 'Четверг'
        mes = RASPISANIE(number, day, Name_day)
        msg_out(call.message, mes)
    elif call.data == "friday":
        day = 4
        Name_day = 'Пятница'
        mes = RASPISANIE(number, day, Name_day)
        msg_out(call.message, mes)

@bot.message_handler(content_types=['text'])
def callback_worker(message):
    if message.text == "Расписание":
        day_number(message)
    elif message.text == "Выбор группы":
        group_command(message)
    elif message.text == "Помощь":
        help_command(message)


bot.polling(none_stop=True, interval=0)

if __name__ == '__main__':
    bot.polling(none_stop=True)